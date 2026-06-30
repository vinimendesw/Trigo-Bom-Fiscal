"""
Camada 1 + Camada 7 — Exportação para XLSX.
Verifica que o arquivo é gerado com estrutura correta (não apenas sem erro).
"""
import json
import os
import pytest
from openpyxl import load_workbook
import db.repositorio as repo
from exportacao.ordem_compra_xlsx import exportar_itens_oc


OC_COM_ITENS = {
    "numero": "OC-XLSX-001",
    "fornecedor": "Fornecedor Teste",
    "data_emissao": "2026-01-10",
    "itens": [
        {"descricao": "Açúcar 1kg", "unidade": "UN", "quantidade": 50, "valor_unitario": 4.5, "valor_total": 225.0},
        {"descricao": "Sal refinado 1kg", "unidade": "UN", "quantidade": 30, "valor_unitario": 2.0, "valor_total": 60.0},
        {"descricao": "Óleo de soja 900ml", "unidade": "UN", "quantidade": 25, "valor_unitario": 7.0, "valor_total": 175.0},
    ],
}


def test_arquivo_xlsx_criado(db_isolado, tmp_path):
    oc_id = json.loads(repo.salvar_ordem_compra(json.dumps(OC_COM_ITENS)))["id"]
    destino = str(tmp_path / "itens_oc.xlsx")

    resultado = json.loads(exportar_itens_oc(oc_id, destino))

    assert resultado["ok"] is True
    assert os.path.exists(destino)


def test_xlsx_tem_cabecalho_correto(db_isolado, tmp_path):
    oc_id = json.loads(repo.salvar_ordem_compra(json.dumps(OC_COM_ITENS)))["id"]
    destino = str(tmp_path / "itens_oc.xlsx")
    exportar_itens_oc(oc_id, destino)

    wb = load_workbook(destino)
    ws = wb.active
    cabecalho = [ws.cell(1, c).value for c in range(1, 7)]
    assert cabecalho == ["#", "Descrição", "UN", "Quantidade", "Valor Unitário", "Valor Total"]


def test_xlsx_tem_quantidade_correta_de_linhas(db_isolado, tmp_path):
    oc_id = json.loads(repo.salvar_ordem_compra(json.dumps(OC_COM_ITENS)))["id"]
    destino = str(tmp_path / "itens_oc.xlsx")
    exportar_itens_oc(oc_id, destino)

    wb = load_workbook(destino)
    ws = wb.active
    # 1 linha de cabeçalho + 3 itens
    assert ws.max_row == 4


def test_xlsx_valores_numericos_corretos(db_isolado, tmp_path):
    oc_id = json.loads(repo.salvar_ordem_compra(json.dumps(OC_COM_ITENS)))["id"]
    destino = str(tmp_path / "itens_oc.xlsx")
    exportar_itens_oc(oc_id, destino)

    wb = load_workbook(destino)
    ws = wb.active
    # Linha 2 = primeiro item (Açúcar)
    assert ws.cell(2, 1).value == 1           # numeração sequencial
    assert ws.cell(2, 2).value == "Açúcar 1kg"
    assert ws.cell(2, 3).value == "UN"
    assert ws.cell(2, 4).value == pytest.approx(50)
    assert ws.cell(2, 5).value == pytest.approx(4.5)
    assert ws.cell(2, 6).value == pytest.approx(225.0)


def test_xlsx_oc_sem_itens_gera_apenas_cabecalho(db_isolado, tmp_path):
    oc_id = json.loads(repo.salvar_ordem_compra(json.dumps({**OC_COM_ITENS, "itens": []})))["id"]
    destino = str(tmp_path / "vazio.xlsx")
    exportar_itens_oc(oc_id, destino)

    wb = load_workbook(destino)
    ws = wb.active
    assert ws.max_row == 1  # só o cabeçalho
